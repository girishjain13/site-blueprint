"""Renders the HTML report and produces JSON/CSV/XLSX exports from audit data."""
from __future__ import annotations

import csv
import io
import json
from pathlib import Path

from jinja2 import Environment, FileSystemLoader, select_autoescape

TEMPLATE_DIR = Path(__file__).parent / "templates"

_env = Environment(
    loader=FileSystemLoader(str(TEMPLATE_DIR)),
    autoescape=select_autoescape(["html"]),
)


def render_html_report(audit_data: dict, graph_cap: int = 250) -> str:
    template = _env.get_template("report.html")
    return template.render(data=audit_data, data_json=json.dumps(audit_data), graph_cap=graph_cap)


def export_json(audit_data: dict) -> bytes:
    return json.dumps(audit_data, indent=2).encode("utf-8")


def export_csv(audit_data: dict) -> bytes:
    buf = io.StringIO()
    fieldnames = [
        "url", "status_code", "title", "word_count", "path_depth", "click_depth",
        "is_thin_content", "is_duplicate_of", "images_total", "images_missing_alt",
        "has_schema_org", "canonical", "internal_links_out_count", "error",
    ]
    writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in audit_data["pages"].values():
        writer.writerow(row)
    return buf.getvalue().encode("utf-8")


def export_xlsx(audit_data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()

    ws = wb.active
    ws.title = "Page Inventory"
    fieldnames = [
        "url", "status_code", "title", "word_count", "path_depth", "click_depth",
        "is_thin_content", "is_duplicate_of", "images_total", "images_missing_alt",
        "has_schema_org", "canonical", "internal_links_out_count", "error",
    ]
    ws.append(fieldnames)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in audit_data["pages"].values():
        ws.append([row.get(f) for f in fieldnames])

    ws2 = wb.create_sheet("Scores")
    ws2.append(["Metric", "Score"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for k, v in audit_data["scoring"].items():
        if isinstance(v, (int, float, str)):
            ws2.append([k, v])

    ws3 = wb.create_sheet("Action Plan")
    ws3.append(["Priority", "Area", "Action"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    for item in audit_data["scoring"]["action_plan"]:
        ws3.append([item["priority"], item["area"], item["action"]])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
